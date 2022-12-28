from random import random
import openpyxl
# Give the location of the file
my_path = "C:/CyBorG/IIT Kharagpur/Sudoku_Cancer/Based_onJhForrest_Comment/Generation of Sudoku Instances/2) Filtering Solved Sudoku for Problem Creation/Sudoku_Solutions (till 8).xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)

sheets = my_wb_obj.sheetnames

for sheet in sheets:
    my_sheet_obj = my_wb_obj[sheet]

    Sudoku_size=int(sheet)
    print(Sudoku_size)
    print(type(Sudoku_size))

    VALUE=Sudoku_size*Sudoku_size
    for row in range(1,VALUE+1):
        for column in range(1,VALUE+1):
            if 13*random()<11:
            #if 7*random()>5:
                cell_position_in_sheet = my_sheet_obj.cell(row=row, column=column)
                cell_position_in_sheet.value = ""

    my_wb_obj.save("Sudoku_Filtered.xlsx")  # Save the file