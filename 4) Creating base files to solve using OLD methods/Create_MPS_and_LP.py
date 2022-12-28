#https://github.com/coin-or/pulp/blob/master/examples/Sudoku1.py


"""
The Sudoku Problem Formulation for the PuLP Modeller
Authors: Antony Phillips, Dr Stuart Mitchell
edited by Nathan Sudermann-Merx
"""

from math import pow
from openpyxl.styles import Border, Side
from openpyxl import Workbook,load_workbook
from gurobipy import *

# Give the location of the file
Path = "C:/CyBorG/IIT Kharagpur/Sudoku_Cancer/Based_onJhForrest_Comment/Generation of Sudoku Instances/3) Creating base files to solve using OLD methods/Sudoku_Filtered.xlsx"
my_wb_obj = load_workbook(Path)
sheets = my_wb_obj.sheetnames

for sheet in sheets:
    my_sheet_obj = my_wb_obj[sheet]

    Sudoku_size=int(sheet)
    ROWS = COLS = VALUES = range(1, Sudoku_size*Sudoku_size+1)

    # The boxes list is created, with the row and column index of each square in each box
    Boxes = [[(Sudoku_size * i + k + 1, Sudoku_size * j + l + 1) for k in range(Sudoku_size) for l in range(Sudoku_size)] for i in range(Sudoku_size) for j in range(Sudoku_size)]
    #print(Boxes)
    # The prob variable is created to contain the problem data
    
    prob = Model("Sudoku Problem")

    # The decision variables are created
    choices={}
    for r in ROWS:
        for c in COLS:
            for val in VALUES:
                choices[r,c,val]=prob.addVar(vtype = GRB.BINARY, name = "x_%d_%d_%d"% (r,c,val))

    # We do not define an objective function since none is needed

    # The row, column and box constraints are added for each value
    prob.addConstrs(quicksum(choices[r,c,val] for val in VALUES) == 1 for r in ROWS for c in COLS)
    prob.addConstrs(quicksum(choices[r,c,val] for r in ROWS) == 1 for val in VALUES for c in COLS)
    prob.addConstrs(quicksum(choices[r,c,val] for c in COLS) == 1 for r in ROWS for val in VALUES)
    for val in VALUES:
        for b in Boxes:
            prob.addConstr(quicksum(choices[r,c,val] for r,c in b) == 1)

    given_Values={}
    for row in ROWS:
        for column in COLS:
            cell_position_in_sheet = my_sheet_obj.cell(row=row, column=column)
            if cell_position_in_sheet.value:
                prob.addConstr(choices[row,column,int(cell_position_in_sheet.value)] == 1)

    # The problem is solved using PuLP's choice of Solver
    prob.write("Filtered_Sudoku_Size"+str(Sudoku_size)+".mps")
    prob.write("Filtered_Sudoku_Size"+str(Sudoku_size)+".lp")
    #prob.optimize()
    # This is for comparing whether Bucket Variables are better/faster for formulating the same problem to be solved efficiently
