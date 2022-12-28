#https://github.com/coin-or/pulp/blob/master/examples/Sudoku1.py

"""
The Sudoku Problem Formulation for the PuLP Modeller

Authors: Antony Phillips, Dr Stuart Mitchell
edited by Nathan Sudermann-Merx
"""

from openpyxl.styles import Border, Side
from openpyxl import Workbook
from gurobipy import *

wb = Workbook()
del wb["Sheet"]
double_border = Side(border_style="double", color="000000")
tolerance = 1e-5

# All rows, columns and values within a Sudoku take values from 1 to Size of the Sudoku

for Sudoku_size in range(9,33+1):

    VALS = ROWS = COLS = range(1, Sudoku_size*Sudoku_size+1)

    # The boxes list is created, with the row and column index of each square in each box
    Boxes = [[(Sudoku_size * i + k + 1, Sudoku_size * j + l + 1) for k in range(Sudoku_size) for l in range(Sudoku_size)] for i in range(Sudoku_size) for j in range(Sudoku_size)]
    #print(Boxes)
    # The prob variable is created to contain the problem data
    
    prob = Model("Sudoku Problem")

    # The decision variables are created
    choices={}
    for v in VALS:
        for r in ROWS:
            for c in COLS:
                choices[v,r,c]=prob.addVar(vtype = GRB.BINARY, name = "x_%d_%d_%d"% (v,r,c))

    # We do not define an objective function since none is needed

    # A constraint ensuring that only one value can be in each square is created
    prob.addConstrs(quicksum(choices[v,r,c] for v in VALS) == 1 for r in ROWS for c in COLS)

    # The row, column and box constraints are added for each value
    prob.addConstrs(quicksum(choices[v,r,c] for c in COLS) == 1 for r in ROWS for v in VALS)
    prob.addConstrs(quicksum(choices[v,r,c] for r in ROWS) == 1 for c in COLS for v in VALS)

    for v in VALS:
        for b in Boxes:
            prob.addConstr(quicksum(choices[v,r,c] for r,c in b) == 1)
    
    # The problem is solved using PuLP's choice of Solver
    prob.write("Sudoku_Size"+str(Sudoku_size)+".mps")
    prob.write("Sudoku_Size"+str(Sudoku_size)+".lp")
    prob.optimize()
    
    ws = wb.create_sheet(str(Sudoku_size))
    for r in ROWS:
        for c in COLS:
            for v in VALS:
                if choices[v,r,c].x <= 1+tolerance and choices[v,r,c].x >= 1-tolerance:
                    cell_position_in_sheet = ws.cell(row=r, column=c)
                    cell_position_in_sheet.value = v

    for times in range(0,Sudoku_size+1):
        for c in COLS:
            cell_position_in_sheet = ws.cell(row=times*Sudoku_size+1, column=c)
            cell_position_in_sheet.border = Border(top=double_border)
        for r in ROWS:
            cell_position_in_sheet = ws.cell(row=r, column=times*Sudoku_size+1)
            cell_position_in_sheet.border = Border(left=double_border)

    for times_1 in range(0,Sudoku_size):
        for times_2 in range(0,Sudoku_size):
            cell_position_in_sheet = ws.cell(row=times_2*Sudoku_size+1, column=times_1*Sudoku_size+1)
            cell_position_in_sheet.border = Border(left=double_border , top=double_border)